#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IPTV数据批量导入工具
功能：将文本文件（频道列表、测试结果、模板文件、服务器地址）导入到Excel工作表
支持五类导入任务，自动清除旧数据、填充公式、备份原文件
"""

import os
import sys
import io
import threading
import time
import shutil
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# 解决Windows控制台UTF-8乱码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)

# ==================== 配置区域 ====================
REGIONS = [
    "江苏", "上海", "浙江", "广东", "北京", "四川", "山东", "福建", "贵州", "重庆",
    "青海", "广西", "湖南", "宁夏", "云南", "内蒙古", "天津", "安徽", "山西", "江西",
    "河北", "河南", "海南", "湖北", "甘肃", "辽宁", "吉林", "陕西", "黑龙江", "新疆"
]
OPERATORS = ["电信", "移动", "联通"]
VALID_SHEET_NAMES = {region + op for region in REGIONS for op in OPERATORS}

START_ROW = 2   # 所有数据均从第2行开始（第1行作为表头保留）

def select_channel_columns(values):
    """频道列表列选择函数：2列时取全部，3列时取第2、3列，其他情况取前2列"""
    if len(values) == 2:
        return values
    if len(values) == 3:
        return [values[1], values[2]]
    return values[:2]

# 定义五类导入任务
IMPORT_TASKS = [
    {
        "suffix": ".txt",
        "start_col": "A",
        "end_col": "B",
        "delimiter": ",",
        "description": "频道列表",
        "select_columns": select_channel_columns
    },
    {
        "suffix": "_quick.txt",
        "start_col": "S",
        "end_col": "U",
        "delimiter": "\t",
        "description": "快速测试",
        "select_columns": None
    },
    {
        "suffix": "_probe.txt",
        "start_col": "W",
        "end_col": "AB",
        "delimiter": "\t",
        "description": "精确测试",
        "select_columns": None
    },
    {
        "description": "模板文件",
        "template": True
    },
    {
        "description": "测试服务器地址",
        "server_address": True
    }
]

# ========== 路径配置 ==========
EXCEL_FILE_NAME = "【IPTV_multicast】.xlsx"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
drive_root = os.path.splitdrive(SCRIPT_DIR)[0] + os.sep
EXCEL_FILE = os.path.join(drive_root, "【IPTV_data】", EXCEL_FILE_NAME)
TEXT_DIR = os.path.join(SCRIPT_DIR, "rtp")
IP_DIR = os.path.join(SCRIPT_DIR, "ip")
SLOW_DIR = os.path.join(IP_DIR, "slow")
TEMPLATE_DIR = os.path.join(SCRIPT_DIR, "template")

# ========== 交互选择函数 ==========
def select_import_mode():
    """选择导入模式，返回任务索引列表或None"""
    print("\n请选择导入模式：")
    print("  1 - 仅导入频道列表（.txt）")
    print("  2 - 仅导入快速测试（_quick.txt）")
    print("  3 - 仅导入精确测试（_probe.txt）")
    print("  4 - 仅导入模板文件（template_.txt）")
    print("  5 - 仅导入测试服务器地址")
    print("  all 或 直接回车 - 导入全部五类")
    print("  q / quit - 退出")
    while True:
        choice = input("请输入数字(1/2/3/4/5)或'all/q'（回车默认全部）: ").strip().lower()
        if choice == '' or choice == 'all':
            return [0, 1, 2, 3, 4]
        if choice == '1':
            return [0]
        if choice == '2':
            return [1]
        if choice == '3':
            return [2]
        if choice == '4':
            return [3]
        if choice == '5':
            return [4]
        if choice in ('q', 'quit'):
            return None
        print("输入无效，请重新输入")

def select_sheets(wb):
    """选择要处理的工作表，返回名称列表或None（表示全部）"""
    valid_sheets = [name for name in wb.sheetnames if name in VALID_SHEET_NAMES]
    if not valid_sheets:
        print("没有符合规则的工作表")
        return None

    print("\n可选工作表列表（按序号或名称选择）：")
    cols = 5
    for i in range(0, len(valid_sheets), cols):
        row_sheets = valid_sheets[i:i+cols]
        row_text = "\t"
        for j, sheet_name in enumerate(row_sheets):
            idx = i + j + 1
            row_text += f"{idx:2d} - {sheet_name}\t"
        print(row_text)

    choice = input("请输入工作表序号或名称（回车默认全部）: ").strip()
    if not choice:
        return None
    if choice.isdigit():
        idx = int(choice) - 1
        if 0 <= idx < len(valid_sheets):
            return [valid_sheets[idx]]
        print(f"序号 {choice} 超出范围")
        return select_sheets(wb)
    if choice in valid_sheets:
        return [choice]
    print(f"未找到工作表：{choice}")
    return select_sheets(wb)

# ========== 功能函数 ==========
def clear_column_range(ws, start_col_idx, end_col_idx, start_row=2):
    """清除指定列区域从start_row开始的所有数据"""
    max_row = ws.max_row
    if max_row < start_row:
        return
    for row in range(start_row, max_row + 1):
        for col in range(start_col_idx, end_col_idx + 1):
            ws.cell(row=row, column=col).value = None

def import_text_to_columns(ws, text_file_path, start_col, end_col, start_row, delimiter, select_columns_func=None):
    """将文本文件内容导入到Excel的指定列区域，返回导入的行数（-1表示文件不存在，0表示空文件）"""
    if not os.path.isfile(text_file_path):
        return -1

    with open(text_file_path, 'r', encoding='utf-8') as f:
        lines = [line.rstrip('\n') for line in f if line.strip() != '']

    if not lines:
        return 0

    start_col_idx = column_index_from_string(start_col)
    end_col_idx = column_index_from_string(end_col)
    max_cols = end_col_idx - start_col_idx + 1

    clear_column_range(ws, start_col_idx, end_col_idx, start_row)

    rows_imported = 0
    for row_offset, line in enumerate(lines):
        values = line.split(delimiter)
        if select_columns_func:
            values = select_columns_func(values)
        if len(values) > max_cols:
            print(f"    警告：第 {row_offset+1} 行有 {len(values)} 列，超出 {max_cols} 列范围，将截断")
            values = values[:max_cols]
        target_row = start_row + row_offset
        for col_offset, val in enumerate(values):
            ws.cell(row=target_row, column=start_col_idx + col_offset, value=val)
        rows_imported += 1
    return rows_imported

def extract_first_valid_server(file_path):
    """
    从测试结果文件中提取第一个有效的服务器地址（去除http://前缀）
    文件格式：服务器地址\t速度\t...
    速度不等于 [X] 的为有效服务器
    """
    if not os.path.exists(file_path):
        return None
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                parts = line.split('\t')
                if len(parts) >= 2:
                    server = parts[0].strip()
                    speed = parts[1].strip()
                    # 跳过无效服务器
                    if speed == '[X]':
                        continue
                    server = server.replace('http://', '').replace('https://', '')
                    if ':' in server:
                        return server
    except Exception:
        pass
    return None

def read_server_address(city_name):
    """
    读取测试服务器地址，按优先级：
    1. ip/{city}_ip_good.txt 中的第一条有效服务器
    2. ip/{city}_ip_precise.txt 中的第一条有效服务器
    3. ip/slow/{city}_ip_precise_slow.txt 中的第一条有效服务器
    4. 以上都没有，返回 "ipipip"
    """
    # 1. 优先从 _ip_good.txt 读取
    good_file = os.path.join(IP_DIR, f"{city_name}_ip_good.txt")
    server = extract_first_valid_server(good_file)
    if server:
        print(f"    从 {city_name}_ip_good.txt 获取服务器: {server}")
        return server

    # 2. 其次从 _ip_precise.txt 读取
    precise_file = os.path.join(IP_DIR, f"{city_name}_ip_precise.txt")
    server = extract_first_valid_server(precise_file)
    if server:
        print(f"    从 {city_name}_ip_precise.txt 获取服务器: {server}")
        return server

    # 3. 最后从 slow 目录下的 _ip_precise_slow.txt 读取
    slow_file = os.path.join(SLOW_DIR, f"{city_name}_ip_precise_slow.txt")
    server = extract_first_valid_server(slow_file)
    if server:
        print(f"    从 slow/{city_name}_ip_precise_slow.txt 获取服务器: {server}")
        return server

    # 4. 默认统配符
    print(f"    未找到任何服务器文件，使用默认占位符: ipipip")
    return "ipipip"

def import_server_address(ws, city_name):
    """导入测试服务器地址到O1单元格"""
    server_addr = read_server_address(city_name)
    o1_cell = ws['O1']
    current_value = str(o1_cell.value) if o1_cell.value else ""

    if "测试地址:" not in current_value:
        print(f"  ⚠ O1单元格格式不正确，未导入")
        return False

    start_idx = current_value.find("测试地址:") + 5
    if "{" in current_value and "}" in current_value:
        start_brace = current_value.find("{", start_idx)
        end_brace = current_value.find("}", start_brace)
        if start_brace != -1 and end_brace != -1:
            new_value = (current_value[:start_brace] + "{" + server_addr + "}" +
                         current_value[end_brace+1:])
        else:
            new_value = current_value[:start_idx] + "{" + server_addr + "}"
    else:
        new_value = current_value[:start_idx] + "{" + server_addr + "}"
    o1_cell.value = new_value
    print(f"  ✓ 测试服务器地址导入成功: {server_addr}")
    return True

def import_template_file(ws, city_name):
    """导入模板文件（template_{city}.txt）到AD-AE列"""
    template_file = os.path.join(TEMPLATE_DIR, f"template_{city_name}.txt")
    if not os.path.isfile(template_file):
        print(f"  ✗ 模板文件不存在：template_{city_name}.txt")
        return False

    start_col_idx = column_index_from_string("AD")
    end_col_idx = column_index_from_string("AE")
    clear_column_range(ws, start_col_idx, end_col_idx, START_ROW)

    try:
        with open(template_file, 'r', encoding='utf-8') as f:
            lines = [line.rstrip('\n') for line in f]
    except Exception as e:
        print(f"  ✗ 读取模板文件失败: {e}")
        return False

    rows_imported = 0
    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith('#'):
            continue
        if '\t' in stripped:
            parts = stripped.split('\t', 1)
        else:
            parts = stripped.split(',', 1)
        if len(parts) >= 2:
            name = parts[0].strip()
            url = parts[1].strip()
            target_row = START_ROW + rows_imported
            ws.cell(row=target_row, column=start_col_idx, value=name)
            ws.cell(row=target_row, column=start_col_idx + 1, value=url)
            rows_imported += 1

    if rows_imported > 0:
        print(f"  ✓ 模板文件导入 {rows_imported} 行数据（列 AD-AE）")
        return True
    print(f"  ⚠ 模板文件为空或无有效数据")
    return False

def fill_formulas(ws, num_rows):
    """复制第2行的完整格式到目标行，包含公式和样式"""
    if num_rows <= 1:
        print(f"  ✓ 公式填充完成，共填充 0 行")
        return True

    columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    source_formulas = []
    source_styles = []

    for col_idx in range(len(columns)):
        source_cell = ws.cell(row=2, column=col_idx + 3)
        cell_value = source_cell.value

        if hasattr(cell_value, 'text'):
            formula_text = cell_value.text
        elif hasattr(cell_value, 'formula'):
            formula_text = cell_value.formula
        elif isinstance(cell_value, str):
            formula_text = cell_value
        else:
            formula_text = None

        if formula_text and isinstance(formula_text, str):
            stripped = formula_text.strip()
            if stripped.startswith('=') or (stripped.startswith('{=') and stripped.endswith('}')):
                source_formulas.append(stripped)
            else:
                source_formulas.append(None)
        else:
            source_formulas.append(None)

        source_styles.append({
            'font': copy(source_cell.font) if source_cell.font else None,
            'fill': copy(source_cell.fill) if source_cell.fill else None,
            'border': copy(source_cell.border) if source_cell.border else None,
            'alignment': copy(source_cell.alignment) if source_cell.alignment else None,
            'number_format': source_cell.number_format,
            'protection': copy(source_cell.protection) if source_cell.protection else None
        })

    filled_count = 0
    for row_offset in range(num_rows - 1):
        current_row = 3 + row_offset
        for col_idx, formula in enumerate(source_formulas):
            if formula:
                if formula.startswith('{=') and formula.endswith('}'):
                    inner = formula[2:-1]
                    new_inner = inner.replace('A2', f'A{current_row}').replace('B2', f'B{current_row}')
                    new_formula = f'{{={new_inner}}}'
                else:
                    new_formula = formula.replace('A2', f'A{current_row}').replace('B2', f'B{current_row}')
                target_cell = ws.cell(row=current_row, column=col_idx + 3)
                target_cell.value = new_formula
                style = source_styles[col_idx]
                if style['font']:
                    target_cell.font = copy(style['font'])
                if style['fill']:
                    target_cell.fill = copy(style['fill'])
                if style['border']:
                    target_cell.border = copy(style['border'])
                if style['alignment']:
                    target_cell.alignment = copy(style['alignment'])
                if style['number_format']:
                    target_cell.number_format = style['number_format']
                if style['protection']:
                    target_cell.protection = copy(style['protection'])
                filled_count += 1

    print(f"  ✓ 公式填充完成，共填充 {filled_count} 个公式，已复制格式")
    return True

def batch_import(selected_task_indices, wb, selected_sheets=None):
    """批量导入主函数"""
    if not os.path.isfile(EXCEL_FILE):
        print(f"错误：Excel 文件不存在 - {EXCEL_FILE}")
        return
    if not os.path.isdir(TEXT_DIR):
        print(f"错误：文本目录不存在 - {TEXT_DIR}")
        return

    print(f"\n已打开文件：{os.path.basename(EXCEL_FILE)}")
    print(f"文本目录：{TEXT_DIR}")
    tasks_to_run = [IMPORT_TASKS[i] for i in selected_task_indices]
    print("本次将导入：", ", ".join(t['description'] for t in tasks_to_run))

    all_valid_sheets = [name for name in wb.sheetnames if name in VALID_SHEET_NAMES]
    if selected_sheets is not None:
        sheets_to_process = selected_sheets
        print(f"待处理工作表：{len(sheets_to_process)} 个（{', '.join(sheets_to_process)}）")
    else:
        sheets_to_process = all_valid_sheets
        print(f"待处理工作表：{len(sheets_to_process)} 个")
    print()

    total_success_sheets = 0
    skipped_count = len(wb.sheetnames) - len(all_valid_sheets)

    for sheet_name in sheets_to_process:
        print(f"► 处理工作表：{sheet_name}")
        ws = wb[sheet_name]
        sheet_success = False

        for task in tasks_to_run:
            if task.get('server_address'):
                if import_server_address(ws, sheet_name):
                    sheet_success = True
            elif task.get('template'):
                if import_template_file(ws, sheet_name):
                    sheet_success = True
            else:
                txt_filename = f"{sheet_name}{task['suffix']}"
                txt_path = os.path.join(TEXT_DIR, txt_filename)
                rows = import_text_to_columns(
                    ws, txt_path, task['start_col'], task['end_col'],
                    START_ROW, task['delimiter'], task.get('select_columns')
                )
                if rows == -1:
                    print(f"  ✗ {task['description']} 文件不存在：{txt_filename}")
                elif rows == 0:
                    print(f"  ⚠ {task['description']} 文件为空，已清除对应区域数据")
                    sheet_success = True
                else:
                    print(f"  ✓ {task['description']} 导入 {rows} 行数据（列 {task['start_col']}-{task['end_col']}）")
                    sheet_success = True
                    if task['description'] == "频道列表":
                        fill_formulas(ws, rows)

        if sheet_success:
            total_success_sheets += 1
        print()

    if skipped_count > 0 and selected_sheets is None:
        print(f"跳过不符合规则的工作表：{skipped_count} 个")

    if total_success_sheets == 0:
        print("\n⚠️ 没有成功处理任何数据，文件未保存")
        return

    print(f"\n正在保存文件：{EXCEL_FILE}...")
    name, ext = os.path.splitext(EXCEL_FILE)
    backup_file = f"{name}_backup{ext}"
    print(f"  创建备份文件：{backup_file}")
    try:
        shutil.copy2(EXCEL_FILE, backup_file)
        print(f"  ✅ 备份文件创建成功")
    except Exception as e:
        print(f"  ⚠️ 备份文件创建失败: {e}")

    def attempt_save():
        try:
            wb.save(EXCEL_FILE)
            print(f"✅ 已保存文件：{EXCEL_FILE}")
            return True
        except PermissionError:
            return False

    if attempt_save():
        return

    while True:
        print(f"\n⚠️ 无法保存原文件（文件可能被其他程序占用）")
        choice = input("请选择：\n  1 - 继续尝试覆盖保存\n  2 - 另存为带日期时间后缀的新文件\n请输入选择(1/2) [默认1]: ").strip()
        if choice == '' or choice == '1':
            print("  正在重试保存...")
            if attempt_save():
                break
        elif choice == '2':
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_file = f"{name}_updated_{timestamp}{ext}"
            wb.save(new_file)
            print(f"✅ 已另存为新文件：{new_file}")
            break
        else:
            print("输入无效，请重新输入（1/2）")

def load_workbook_async(excel_file, result_dict):
    """异步加载Excel工作簿"""
    try:
        wb = load_workbook(excel_file)
        result_dict['wb'] = wb
        result_dict['error'] = None
    except Exception as e:
        result_dict['wb'] = None
        result_dict['error'] = str(e)

# ==================== 主函数 ====================
def main():
    print("=" * 60)
    print("IPTV数据批量导入工具")
    print("功能：导入频道列表、测试结果、模板文件到Excel工作表")
    print("=" * 60)
    print(f"脚本目录: {SCRIPT_DIR}")
    print(f"盘符根目录: {drive_root}")
    print(f"Excel 文件: {EXCEL_FILE}")
    print(f"文本目录: {TEXT_DIR}")

    result_dict = {}
    thread = threading.Thread(target=load_workbook_async, args=(EXCEL_FILE, result_dict), daemon=True)
    thread.start()

    selected_tasks = select_import_mode()
    if selected_tasks is None:
        print("\n退出程序")
        sys.exit(0)

    if 'wb' not in result_dict:
        print(f"\n正在打开文件：{EXCEL_FILE}...")
        while 'wb' not in result_dict:
            time.sleep(0.1)

    if result_dict.get('error'):
        print(f"错误：无法打开文件 - {result_dict['error']}")
        sys.exit(1)

    wb = result_dict['wb']
    selected_sheets = select_sheets(wb)
    batch_import(selected_tasks, wb, selected_sheets)

if __name__ == "__main__":
    main()