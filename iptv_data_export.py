#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IPTV数据批量导出工具
功能：从Excel工作表导出数据到文本文件
支持五类导出任务，自动备份已存在的文件，支持连续导出多个城市
新增：定制模板导出功能（根据Q列标记导出选中的频道）
改进：优化Excel读取，支持更多格式
"""

import os
import sys
import io
import threading
import time
import shutil
import re
from datetime import datetime
from pathlib import Path

# 尝试导入openpyxl，如果失败则尝试xlrd/xlwt
try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
    USE_OPENPYXL = True
except ImportError:
    USE_OPENPYXL = False
    try:
        import xlrd
        import xlwt
    except ImportError:
        print("错误: 请安装 openpyxl 或 xlrd/xlwt")
        print("推荐安装: pip install openpyxl")
        sys.exit(1)

# 解决Windows控制台UTF-8乱码
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)

# ==================== 配置区域 ====================
class ExportConfig:
    """导出配置类"""
    # 支持的地区和运营商
    REGIONS = [
        "江苏", "上海", "浙江", "广东", "北京", "四川", "山东", "福建", "贵州", "重庆",
        "青海", "广西", "湖南", "宁夏", "云南", "内蒙古", "天津", "安徽", "山西", "江西",
        "河北", "河南", "海南", "湖北", "甘肃", "辽宁", "吉林", "陕西", "黑龙江", "新疆"
    ]
    OPERATORS = ["电信", "移动", "联通"]
    
    # Excel 配置
    EXCEL_DIR_NAME = "【IPTV_data】"
    EXCEL_FILE_NAME = "【IPTV_multicast】.xlsx"
    START_ROW = 2  # 数据从第2行开始（第1行作为表头）
    
    # 模板文件配置（从AD-AE列导出）
    TEMPLATE_CONFIG = {
        "output_dir": "template",           # 输出到 template 目录
        "backup_dir": "template/backup",    # 备份到 template/backup
        "filename_prefix": "template_",     # 文件名前缀
        "need_backup": True,                # 需要备份
    }
    
    # 定制模板配置（根据Q列标记导出）
    CUSTOM_TEMPLATE_CONFIG = {
        "enabled": True,
        "marker_column": "Q",           # 标记列（√ 或 是 表示选入）
        "channel_name_col": "A",        # 频道名列
        "channel_url_col": "B",         # 频道组播源列
        "default_protocol": "rtp",      # 默认协议（rtp 或 udp）
        "server_placeholder": "ipipip", # 服务器占位符
        "output_dir": "template/export", # 输出到 template/export 目录
        "filename_prefix": "template_",  # 文件名前缀（与模板文件同名）
        "need_backup": False,            # 不需要备份
        "include_header": True,         # 是否包含表头
        "header_text": "定制模板（精选频道）",  # 表头文本
        "comment_prefix": "#",          # 注释前缀
    }
    
    # 导出任务配置
    EXPORT_TASKS = [
        {
            "name": "频道列表",
            "suffix": ".txt",
            "start_col": "A",
            "end_col": "B",
            "output_subdir": "rtp",
            "backup_subdir": "rtp/backup",
            "delimiter": ",",
            "needs_sort": True,
            "need_backup": True,
        },
        {
            "name": "快速测试",
            "suffix": "_checked.txt",
            "start_col": "S",
            "end_col": "U",
            "output_subdir": "rtp",
            "backup_subdir": "rtp/backup",
            "delimiter": "\t",
            "needs_sort": False,
            "need_backup": True,
        },
        {
            "name": "精确测试",
            "suffix": "_source.txt",
            "start_col": "W",
            "end_col": "AB",
            "output_subdir": "rtp",
            "backup_subdir": "rtp/backup",
            "delimiter": "\t",
            "needs_sort": False,
            "need_backup": True,
        },
        {
            "name": "模板文件",
            "start_col": "AD",
            "end_col": "AE",
            "is_template": True,  # 标记为模板文件类型
        },
        {
            "name": "定制模板",
            "is_custom_template": True,  # 标记为定制模板类型
        }
    ]

config = ExportConfig()

# ========== 工具函数 ==========
def is_valid_sheet_name(sheet_name):
    """
    判断工作表名称是否有效
    规则：包含地区+运营商即可（不要求完全匹配）
    例如：广东电信中信、广东电信华为、广东电信_测试 等都有效
    """
    # 检查是否包含任意地区
    has_region = any(region in sheet_name for region in config.REGIONS)
    # 检查是否包含任意运营商
    has_operator = any(operator in sheet_name for operator in config.OPERATORS)
    
    return has_region and has_operator

def natural_sort_key(channel_line):
    """
    自然排序键函数
    将字符串中的数字部分转换为整数进行排序
    例如：CCTV1, CCTV2, CCTV10 会按数字大小排序
    """
    channel_name = channel_line.split(',')[0].strip()
    parts = re.split(r'(\d+)', channel_name)
    result = []
    for part in parts:
        if part.isdigit():
            result.append(int(part))
        else:
            result.append(part)
    return result

def get_excel_file_path():
    """
    获取Excel文件路径
    规则：在项目所在盘的根目录下的 【IPTV_data】 文件夹中
    例如：I:/【IPTV_data】/【IPTV_multicast】.xlsx
    """
    # 获取当前脚本所在盘的根目录
    script_path = Path(__file__).absolute()
    drive_root = Path(script_path.drive + os.sep)  # 例如：I:\
    
    # 构建Excel文件路径：盘符:/【IPTV_data】/【IPTV_multicast】.xlsx
    excel_path = drive_root / config.EXCEL_DIR_NAME / config.EXCEL_FILE_NAME
    
    return excel_path

def get_output_base_dir():
    """
    获取输出基础目录
    在脚本所在目录下创建输出文件夹
    """
    return Path(__file__).parent.absolute()

def backup_existing_file(file_path, backup_dir):
    """
    如果文件已存在，将其移动到对应的 backup 目录
    返回 (是否备份, 备份文件路径)
    """
    if not os.path.exists(file_path):
        return False, None
    
    os.makedirs(backup_dir, exist_ok=True)
    
    filename = os.path.basename(file_path)
    backup_path = os.path.join(backup_dir, filename)
    
    try:
        if os.path.exists(backup_path):
            os.remove(backup_path)
        shutil.move(file_path, backup_path)
        return True, backup_path
    except Exception as e:
        print(f"    备份失败: {e}")
        return False, None

def safe_write_file(file_path, content_lines, backup_dir=None, need_backup=True):
    """
    安全写入文件
    如果 need_backup=True 且 backup_dir 存在，则先备份原文件，再写入新内容
    返回 (是否备份, 备份文件路径)
    """
    backed_up = False
    backup_path = None
    
    if need_backup and backup_dir:
        backed_up, backup_path = backup_existing_file(file_path, backup_dir)
    
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    with open(file_path, 'w', encoding='utf-8') as f:
        for line in content_lines:
            f.write(line + '\n')
    
    return backed_up, backup_path

def read_excel_cell_value(cell):
    """安全读取Excel单元格值"""
    if cell is None:
        return ""
    value = cell.value
    if value is None:
        return ""
    return str(value).strip()

# ========== 导出功能函数 ==========
def export_column_range(ws, task, sheet_name, base_dir):
    """
    导出Excel列区域到文本文件
    返回 (导出行数, 是否备份, 备份路径, 输出文件路径)
    """
    start_col = task['start_col']
    end_col = task['end_col']
    
    start_col_idx = column_index_from_string(start_col)
    end_col_idx = column_index_from_string(end_col)
    
    data_rows = []
    for row in range(config.START_ROW, ws.max_row + 1):
        row_values = []
        has_data = False
        for col in range(start_col_idx, end_col_idx + 1):
            cell_value = read_excel_cell_value(ws.cell(row=row, column=col))
            if cell_value:
                has_data = True
            row_values.append(cell_value)
        
        if has_data:
            data_rows.append(row_values)
    
    if not data_rows:
        return 0, False, None, None
    
    # 生成输出行
    output_lines = []
    for row in data_rows:
        line = task['delimiter'].join(row)
        output_lines.append(line)
    
    # 排序（如果需要）
    if task.get('needs_sort', False):
        output_lines.sort(key=natural_sort_key)
    
    # 生成输出文件路径
    output_dir = os.path.join(base_dir, task['output_subdir'])
    output_file = os.path.join(output_dir, f"{sheet_name}{task['suffix']}")
    backup_dir = os.path.join(base_dir, task['backup_subdir']) if task.get('need_backup', True) else None
    
    backed_up, backup_path = safe_write_file(output_file, output_lines, backup_dir, task.get('need_backup', True))
    
    return len(data_rows), backed_up, backup_path, output_file

def export_template_file(ws, sheet_name, base_dir):
    """
    从AD-AE列导出模板文件
    输出路径：template/template_数据表名.txt
    备份路径：template/backup/template_数据表名.txt
    """
    start_col = "AD"
    end_col = "AE"
    
    start_col_idx = column_index_from_string(start_col)
    end_col_idx = column_index_from_string(end_col)
    
    output_lines = []
    for row in range(config.START_ROW, ws.max_row + 1):
        name = read_excel_cell_value(ws.cell(row=row, column=start_col_idx))
        url = read_excel_cell_value(ws.cell(row=row, column=start_col_idx + 1))
        
        if name and url:
            output_lines.append(f"{name},{url}")
    
    if not output_lines:
        return 0, False, None, None
    
    # 生成输出文件路径
    output_dir = os.path.join(base_dir, config.TEMPLATE_CONFIG['output_dir'])
    output_file = os.path.join(output_dir, f"{config.TEMPLATE_CONFIG['filename_prefix']}{sheet_name}.txt")
    backup_dir = os.path.join(base_dir, config.TEMPLATE_CONFIG['backup_dir'])
    
    backed_up, backup_path = safe_write_file(
        output_file, output_lines, backup_dir, 
        config.TEMPLATE_CONFIG['need_backup']
    )
    
    return len(output_lines), backed_up, backup_path, output_file

def export_custom_template(ws, sheet_name, base_dir):
    """
    导出定制模板
    根据Q列的标记（√ 或 是）筛选频道
    输出路径：template/export/template_数据表名.txt
    无需备份
    """
    marker_col = column_index_from_string(config.CUSTOM_TEMPLATE_CONFIG['marker_column'])
    name_col = column_index_from_string(config.CUSTOM_TEMPLATE_CONFIG['channel_name_col'])
    url_col = column_index_from_string(config.CUSTOM_TEMPLATE_CONFIG['channel_url_col'])
    
    selected_channels = []
    
    for row in range(config.START_ROW, ws.max_row + 1):
        marker_value = read_excel_cell_value(ws.cell(row=row, column=marker_col))
        
        # 检查是否被选中
        is_selected = False
        if marker_value in ['√', '是', '✔', '✓', 'true', 'True', '1', 'yes', 'Yes']:
            is_selected = True
        elif marker_value.lower() in ['true', 'yes', 'y']:
            is_selected = True
        elif marker_value.isdigit() and int(marker_value) == 1:
            is_selected = True
        
        if is_selected:
            channel_name = read_excel_cell_value(ws.cell(row=row, column=name_col))
            channel_url = read_excel_cell_value(ws.cell(row=row, column=url_col))
            
            if channel_name and channel_url:
                # 处理URL
                url = channel_url.strip()
                if url.startswith('rtp://'):
                    url = url[6:]
                elif url.startswith('udp://'):
                    url = url[6:]
                
                protocol = config.CUSTOM_TEMPLATE_CONFIG['default_protocol']
                output_url = f"http://{config.CUSTOM_TEMPLATE_CONFIG['server_placeholder']}/{protocol}/{url}"
                selected_channels.append(f"{channel_name},{output_url}")
    
    if not selected_channels:
        return 0, False, None, None
    
    # 自然排序
    selected_channels.sort(key=natural_sort_key)
    
    # 生成输出
    output_lines = []
    if config.CUSTOM_TEMPLATE_CONFIG['include_header']:
        header = f"{config.CUSTOM_TEMPLATE_CONFIG['comment_prefix']} {config.CUSTOM_TEMPLATE_CONFIG['header_text']} {datetime.now().strftime('%Y/%m/%d %H:%M')}"
        output_lines.append(header)
        output_lines.append("")
    
    output_lines.extend(selected_channels)
    
    # 生成输出文件路径
    output_dir = os.path.join(base_dir, config.CUSTOM_TEMPLATE_CONFIG['output_dir'])
    output_file = os.path.join(output_dir, f"{config.CUSTOM_TEMPLATE_CONFIG['filename_prefix']}{sheet_name}.txt")
    
    # 定制模板不需要备份
    backed_up, backup_path = safe_write_file(
        output_file, output_lines, None, 
        config.CUSTOM_TEMPLATE_CONFIG['need_backup']
    )
    
    return len(selected_channels), backed_up, backup_path, output_file

def export_single_city(ws, sheet_name, base_dir, selected_tasks):
    """
    导出单个城市的数据
    返回 (成功导出的文件数, 文件列表)
    """
    exported_files = []
    messages = []
    
    for idx, task in enumerate(config.EXPORT_TASKS):
        if selected_tasks != 'all' and idx not in selected_tasks:
            continue
        
        rows = 0
        backed_up = False
        backup_path = None
        output_file = None
        
        try:
            if task.get('is_custom_template', False):
                if config.CUSTOM_TEMPLATE_CONFIG['enabled']:
                    rows, backed_up, backup_path, output_file = export_custom_template(ws, sheet_name, base_dir)
                    if rows > 0:
                        rel_path = os.path.relpath(output_file, base_dir)
                        messages.append(f"  ✓ 定制模板: {rel_path} ({rows}个频道)")
                        exported_files.append(output_file)
                    else:
                        messages.append(f"  ✗ 定制模板: 未选中任何频道")
                else:
                    messages.append(f"  ⊘ 定制模板功能未启用")
                    
            elif task.get('is_template', False):
                rows, backed_up, backup_path, output_file = export_template_file(ws, sheet_name, base_dir)
                if rows > 0:
                    rel_path = os.path.relpath(output_file, base_dir)
                    messages.append(f"  ✓ 模板文件: {rel_path} ({rows}行)")
                    if backed_up:
                        backup_rel = os.path.relpath(backup_path, base_dir)
                        messages.append(f"      备份: {backup_rel}")
                    exported_files.append(output_file)
                else:
                    messages.append(f"  ✗ 模板文件: 无数据")
                    
            else:
                rows, backed_up, backup_path, output_file = export_column_range(ws, task, sheet_name, base_dir)
                if rows > 0:
                    rel_path = os.path.relpath(output_file, base_dir)
                    messages.append(f"  ✓ {task['name']}: {rel_path} ({rows}行)")
                    if backed_up:
                        backup_rel = os.path.relpath(backup_path, base_dir)
                        messages.append(f"      备份: {backup_rel}")
                    exported_files.append(output_file)
                    
        except Exception as e:
            messages.append(f"  ✗ {task['name']}: 导出失败 - {str(e)}")
    
    if messages:
        print(f"\n► {sheet_name}")
        for msg in messages:
            print(msg)
    
    return len(exported_files), exported_files

def list_available_sheets(wb):
    """列出所有可用的工作表（按4列对齐显示）"""
    # 获取所有有效工作表（支持灵活命名）
    valid_sheets = [name for name in wb.sheetnames if is_valid_sheet_name(name)]
    
    if not valid_sheets:
        print("\n⚠ 没有符合规则的工作表")
        print(f"   工作表命名规则: 必须包含地区和运营商")
        print(f"   地区: {', '.join(config.REGIONS[:5])}... (共{len(config.REGIONS)}个)")
        print(f"   运营商: {', '.join(config.OPERATORS)}")
        return []
    
    # 按名称排序
    valid_sheets.sort()
    
    print(f"\n📋 可用的工作表 ({len(valid_sheets)}个):")
    
    # 4列对齐显示
    cols = 4
    # 计算每列的最大宽度（用于对齐）
    col_widths = [0] * cols
    for i, name in enumerate(valid_sheets):
        col_idx = i % cols
        col_widths[col_idx] = max(col_widths[col_idx], len(name) + 5)  # +5 为序号和空格
    
    # 按行输出
    for i in range(0, len(valid_sheets), cols):
        row_items = []
        for j in range(cols):
            if i + j < len(valid_sheets):
                name = valid_sheets[i + j]
                num = i + j + 1
                # 格式化：序号. 工作表名
                item = f"{num:2d}. {name}"
                row_items.append(item.ljust(col_widths[j]))
        print("   " + "".join(row_items))
    
    return valid_sheets

def select_sheets_interactive(wb):
    """交互式选择工作表"""
    valid_sheets = list_available_sheets(wb)
    if not valid_sheets:
        return None
    
    print("\n📌 选择方式:")
    print("   1. 输入序号（如 1）选择单个城市")
    print("   2. 输入多个序号（如 1,3,5）选择多个")
    print("   3. 输入范围（如 1-5）选择连续")
    print("   4. 直接回车选择全部城市")
    print("   5. 输入 q 返回")
    
    choice = input("\n请选择: ").strip()
    
    if choice.lower() == 'q':
        return None
    
    if not choice:
        return valid_sheets
    
    selected = []
    parts = choice.replace(' ', '').split(',')
    for part in parts:
        if '-' in part:
            start, end = part.split('-')
            start_idx = int(start) - 1
            end_idx = int(end) - 1
            if 0 <= start_idx < len(valid_sheets) and 0 <= end_idx < len(valid_sheets):
                for idx in range(start_idx, end_idx + 1):
                    selected.append(valid_sheets[idx])
            else:
                print(f"⚠ 无效范围: {part}")
        elif part.isdigit():
            idx = int(part) - 1
            if 0 <= idx < len(valid_sheets):
                selected.append(valid_sheets[idx])
            else:
                print(f"⚠ 无效序号: {part}")
        else:
            # 支持直接输入工作表名
            if part in valid_sheets:
                selected.append(part)
            else:
                print(f"⚠ 无效工作表名: {part}")
    
    return selected if selected else None

def select_export_mode():
    """选择导出模式"""
    print("\n" + "=" * 50)
    print("导出模式选择")
    print("=" * 50)
    print("  1 - 仅导出频道列表")
    print("  2 - 仅导出快速测试")
    print("  3 - 仅导出精确测试")
    print("  4 - 仅导出模板文件")
    print("  5 - 仅导出定制模板")
    print("  all - 导出全部")
    print("  q - 退出")
    
    while True:
        choice = input("\n请选择: ").strip().lower()
        if choice == 'all':
            return 'all'
        if choice == '1':
            return [0]
        if choice == '2':
            return [1]
        if choice == '3':
            return [2]
        if choice == '4':
            return [3]
        if choice == '5':
            return [4]
        if choice == 'q':
            return None
        print("❌ 输入无效，请输入 1/2/3/4/5/all/q")

def load_workbook_async(excel_file, result_dict):
    """异步加载Excel工作簿"""
    try:
        if USE_OPENPYXL:
            wb = load_workbook(excel_file, data_only=True)
        else:
            # 使用xlrd读取（仅支持.xls）
            wb = xlrd.open_workbook(excel_file)
        result_dict['wb'] = wb
        result_dict['error'] = None
    except Exception as e:
        result_dict['wb'] = None
        result_dict['error'] = str(e)

# ==================== 主函数 ====================
def main():
    print("=" * 60)
    print("📺 IPTV数据批量导出工具 v2.0")
    print("=" * 60)
    
    # 获取Excel文件路径
    excel_file = get_excel_file_path()
    
    # 检查文件是否存在
    if not os.path.isfile(excel_file):
        print(f"\n❌ 错误: Excel文件不存在")
        print(f"   查找路径: {excel_file}")
        print(f"\n请确保文件位于:")
        print(f"   盘符根目录下的 【IPTV_data】 文件夹中")
        print(f"   例如: I:/【IPTV_data】/【IPTV_multicast】.xlsx")
        input("\n按回车键退出...")
        sys.exit(1)
    
    print(f"\n📁 Excel文件: {excel_file}")
    
    # 获取输出基础目录（脚本所在目录）
    base_dir = get_output_base_dir()
    print(f"📁 输出目录: {base_dir}")
    
    # 选择导出模式
    selected_tasks = select_export_mode()
    if selected_tasks is None:
        print("\n👋 退出程序")
        sys.exit(0)
    
    # 加载工作簿
    print(f"\n⏳ 正在加载Excel文件...")
    result_dict = {}
    thread = threading.Thread(target=load_workbook_async, args=(excel_file, result_dict), daemon=True)
    thread.start()
    
    # 等待加载完成
    loading_dots = 0
    while 'wb' not in result_dict:
        loading_dots = (loading_dots + 1) % 4
        sys.stdout.write(f"\r⏳ 加载中{'.' * loading_dots}   ")
        sys.stdout.flush()
        time.sleep(0.2)
    print("\r✅ Excel文件加载完成    ")
    
    if result_dict.get('error'):
        print(f"\n❌ 错误: {result_dict['error']}")
        input("\n按回车键退出...")
        sys.exit(1)
    
    wb = result_dict['wb']
    
    total_exported = 0
    all_exported_files = []
    
    while True:
        selected_sheets = select_sheets_interactive(wb)
        
        if selected_sheets is None:
            again = input("\n是否退出？(y/n): ").strip().lower()
            if again in ('y', 'yes', ''):
                break
            continue
        
        if not selected_sheets:
            continue
        
        print(f"\n{'='*50}")
        print(f"开始导出...")
        print('='*50)
        
        success_count = 0
        for sheet_name in selected_sheets:
            if USE_OPENPYXL:
                ws = wb[sheet_name]
            else:
                ws = wb.sheet_by_name(sheet_name)
            
            files_count, exported_files = export_single_city(ws, sheet_name, base_dir, selected_tasks)
            if files_count > 0:
                success_count += 1
                total_exported += files_count
                all_exported_files.extend(exported_files)
        
        if success_count == 0:
            print(f"\n⚠ 本轮无数据导出")
        else:
            print(f"\n{'='*50}")
            print(f"✅ 本轮完成: {success_count} 个城市, 共导出 {total_exported} 个文件")
        
        # 询问继续
        all_valid = [name for name in wb.sheetnames if is_valid_sheet_name(name)]
        if len(selected_sheets) < len(all_valid):
            again = input("\n继续导出其他城市？(y/n): ").strip().lower()
            if again not in ('y', 'yes'):
                break
        else:
            break
    
    print("\n" + "=" * 60)
    print("✅ 导出完成！")
    print(f"   总计导出: {total_exported} 个文件")
    if all_exported_files:
        print(f"   输出目录: {base_dir}")
        print(f"   输出结构:")
        print(f"     - rtp/                       (频道列表、测试结果)")
        print(f"     - template/                  (模板文件: template_数据表名.txt)")
        print(f"     - template/backup/           (模板文件备份)")
        print(f"     - template/export/           (定制模板: template_数据表名.txt)")
    print("=" * 60)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n👋 用户中断")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        input("\n按回车键退出...")
        sys.exit(1)